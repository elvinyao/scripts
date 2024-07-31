import typer
from fastapi import FastAPI
from jira import JIRA
import pandas as pd
from openpyxl import load_workbook

app = FastAPI()
cli = typer.Typer()

# Jira连接设置
jira = JIRA(server="https://your-jira-instance.com", basic_auth=("username", "password"))

def get_subtasks(key_or_jql):
    if key_or_jql.startswith("project=") or " " in key_or_jql:
        # 假设这是JQL
        issues = jira.search_issues(key_or_jql)
    else:
        # 假设这是一个ticket key
        issues = [jira.issue(key_or_jql)]
    
    subtasks = []
    for issue in issues:
        subtasks.extend(issue.fields.subtasks)
    
    return subtasks

def subtask_to_dict(subtask):
    return {
        "Key": subtask.key,
        "Assignee": subtask.fields.assignee.displayName if subtask.fields.assignee else "Unassigned",
        "Begin Date": str(subtask.fields.created)[:10],
        "End Date": str(subtask.fields.resolutiondate)[:10] if subtask.fields.resolutiondate else "",
        "Summary": subtask.fields.summary
    }

@cli.command()
def export_to_excel(key_or_jql: str, template_path: str, output_path: str):
    subtasks = get_subtasks(key_or_jql)
    data = [subtask_to_dict(subtask) for subtask in subtasks]
    
    # 读取模板Excel文件
    wb = load_workbook(template_path)
    ws = wb.active
    
    # 将数据写入Excel
    for row, item in enumerate(data, start=2):  # 假设第一行是标题
        for col, (key, value) in enumerate(item.items(), start=1):
            ws.cell(row=row, column=col, value=value)
    
    # 保存Excel文件
    wb.save(output_path)
    print(f"Data exported to {output_path}")

if __name__ == "__main__":
    cli()
